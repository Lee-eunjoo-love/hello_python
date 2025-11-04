from selenium import webdriver #. 동적 웹크롤링
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
import os
import time
import openpyxl

#. User-Agent 설정 (서버가 봇을 차단할 수 있으므로 브라우저 정보 추가)
headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.1.1 Safari/605.1.15"}

url = "https://dhlottery.co.kr/gameResult.do?method=byWin"

try:
    #. 데이터 초기화
    lotte_list = []
    title_list = []
    max_number = 1
    
    #. 브라우저 생성
    driver = webdriver.Chrome()
    driver.get(url)
    time.sleep(3)
    
    #. 최근 회차 정보 조회
    dwrNoList_elements = driver.find_element(By.XPATH, '//*[@id="dwrNoList"]/option[1]')    
    dwrNoList_max = dwrNoList_elements.text
    max_number = int(dwrNoList_max) + 1
    
    for i in range(1, max_number):
        #. 각 회차별 브라우저 조회 조건 설정 및 조회 요청
        group_element = driver.find_element(By.ID, "hdrwComb")
        group = Select(group_element)
        selected_group = group.first_selected_option
        if int(i) > 600:
            group.select_by_value('1')
        else:
            group.select_by_value('2')
        dropdown_element = driver.find_element(By.ID, "dwrNoList")
        select = Select(dropdown_element)
        select.select_by_value(str(i))
        search_button = driver.find_element(By.ID, 'searchBtn')
        search_button.send_keys(Keys.RETURN)
        time.sleep(1)

        #. 각 회차별 브라우저 조회 결과 추출
        lotte_title_element = driver.find_element(By.XPATH, '//*[@id="article"]/div[2]/div/div[2]/h4/strong')
        lotte_numbers = []
        lotte_bonus = driver.find_element(By.XPATH, '//*[@id="article"]/div[2]/div/div[2]/div/div[2]/p/span')
        for n in range(6):
            lotte_number_element = driver.find_element(By.XPATH, f'//*[@id="article"]/div[2]/div/div[2]/div/div[1]/p/span[{n + 1}]')
            lotte_numbers.append(lotte_number_element.text)
        lotte_numbers.append(lotte_bonus.text)
        title_list.append(lotte_title_element.text)
        lotte_list.append(lotte_numbers)
        
    #. 엑셀파일 생성 및 활성 시트 불러오기
    book = openpyxl.Workbook()
    sheet = book.active
    
    #. 제목 셀 설정
    sheet.cell(row=1, column=1).value = "회차"
    sheet.cell(row=1, column=2).value = "당첨번호1"
    sheet.cell(row=1, column=3).value = "당첨번호2"
    sheet.cell(row=1, column=4).value = "당첨번호3"
    sheet.cell(row=1, column=5).value = "당첨번호4"
    sheet.cell(row=1, column=6).value = "당첨번호5"
    sheet.cell(row=1, column=7).value = "당첨번호6"
    sheet.cell(row=1, column=8).value = "보너스번호"
       
    #. 데이터 셀 설정
    row = 2
    for t, n in zip(title_list, lotte_list):
        sheet.cell(row=row, column=1).value = t
        for idx, x in enumerate(n):
            sheet.cell(row=row, column = idx + 2).value = x
        row += 1
    
    #. 열 너비 조정
    sheet.column_dimensions["A"].width = 15    
    sheet.column_dimensions["B"].width = 10  
    sheet.column_dimensions["C"].width = 10
    sheet.column_dimensions["D"].width = 10
    sheet.column_dimensions["E"].width = 10
    sheet.column_dimensions["F"].width = 10
    sheet.column_dimensions["G"].width = 10
    sheet.column_dimensions["H"].width = 10
    
    #. 엑셀 파일 저장
    book.save(f'./로또회차별당첨번호.xlsx')
except Exception as e:
    print(f'An error occured: {e}')
    