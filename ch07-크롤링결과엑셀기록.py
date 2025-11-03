import requests
from bs4 import BeautifulSoup
import urllib.request as req
import os
import openpyxl
from openpyxl.drawing.image import Image #. 이미지를 엑셀에 삽입하고 조작
from PIL import Image as PILImage #. pillow 모듈: 이미지 처리

#. User-Agent 설정 (서버가 봇을 차단할 수 있으므로 브라우저 정보 추가)
headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.1.1 Safari/605.1.15"}

#. 박스오피스
url = 'https://search.naver.com/search.naver?sm=tab_sug.top&where=nexearch&ssc=tab.nx.all&query=%ED%98%84%EC%9E%AC%EC%83%81%EC%98%81%EC%98%81%ED%99%94+%EC%88%9C%EC%9C%84&oquery=%ED%98%84%EC%9E%AC%EC%83%81%EC%98%81%EC%98%81%ED%99%94&tqi=jdNzusqVOsVsstJW5qlssssssMw-124645&acq=%ED%98%84%EC%9E%AC%EC%83%81%EC%98%81%EC%98%81%ED%99%94&acr=1&qdt=0&ackey=ookijq85'

#. requests 를 이용해 웹 페이지 HTML 가져오기
try:
    response = requests.get(url, headers=headers)
    response.raise_for_status() #. HTTP 오류 발생시 예외 발생
    
    #. BeautifulSoup 으로 HTML 파싱
    soup = BeautifulSoup(response.text, 'html.parser')
    
    title_elements = soup.select('ul._panel > li strong.name')
    image_elements = soup.select('ul._panel > li img')
    count_elements = soup.select('ul._panel > li span.sub_text')
    
    #. 엑셀 파일 생성
    book = openpyxl.Workbook()
    #. 현재 활성 시트 불러오기
    sheet = book.active
    #. 타이틀
    sheet.cell(row=1, column=1).value = "No"
    sheet.cell(row=1, column=2).value = "영화제목"
    sheet.cell(row=1, column=3).value = "조회수"
    sheet.cell(row=1, column=4).value = "포스터 이미지"
    #. 이미지 다운로드 폴더
    folder_name = './무비차트포스터'
    if not os.path.exists(folder_name):
        os.mkdir(folder_name)
    
    row = 2
    for title, image, count in zip(title_elements, image_elements, count_elements):
        sheet.cell(row=row, column=1).value = row + 1
        sheet.cell(row=row, column=2).value = title.text.strip()
        sheet.cell(row=row, column=3).value = count.text
        img_url = image.attrs['src']
        #. 이미지 크기 조절
        file, ext = os.path.splitext(img_url) #. ext : .을 포함한 확장자
        filename = f'{folder_name}/movie_{row}{ext}'
        req.urlretrieve(img_url, filename)
        img = PILImage.open(filename)
        img_resized = img.resize((int(img.width * 0.3), int(img.height * 0.3)))
        img_resized.save(filename)
        img.close()
        sheet.add_image(Image(filename), f"D{row}") #. 엑셀이 크기 조정된 이미지 첨부        
        # 행 높이 조절
        sheet.row_dimensions[row].height = 87        
        row += 1
        #sheet.cell(row=row, column=4).value = image.attrs['src']
        #print(title.text)
        #print(image.attrs['src'])
        #print(count.text)
        #print('--------------------')
    
    #. 열 너비 조절
    sheet.column_dimensions["A"].width = 23
    sheet.column_dimensions["B"].width = 35
    sheet.column_dimensions["C"].width = 15
    sheet.column_dimensions["D"].width = 21
    #. 엑셀 파일 저장
    book.save('./무비차트.xlsx')
    
except requests.exceptions.RequestException as e:
    print(f'Error during requests to {url}: {e}')
    
except Exception as e:
    print(f'An error occured: {e}')